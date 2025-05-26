import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pdfplumber
import openpyxl
import re
import logging
from datetime import datetime
from openpyxl.styles import Font

logging.getLogger("pdfminer").setLevel(logging.ERROR)

class PDFExportador:
    def __init__(self, root):
        self.root = root
        self.root.title("Exportador de Dados Financeiros - PDF para Excel")
        self.root.geometry("1000x700")
        self.root.configure(bg="#f5f5f5")

        main_frame = tk.Frame(root, bg="#f5f5f5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        campos_frame = tk.LabelFrame(main_frame, text="Campos para Exportação", bg="#f5f5f5", padx=10, pady=10)
        campos_frame.pack(fill=tk.X, padx=5, pady=5)

        self.campos_disponiveis = [
            "Matrícula", "Nome", "CPF", "Banco/Agência", "Conta Corrente",
            "Data Aposentadoria", "Total Rendimentos", "Total Descontos",
            "Depósito FGTS", "Total Líquido", "Margem Consignável 30%", 
            "Margem Consignável 70%", "Bancos Desconto", "Parcelas Desconto", "Valores Desconto"
        ]

        self.campos_selecionados = {campo: tk.BooleanVar(value=True) for campo in self.campos_disponiveis}

        cols = 3
        for i, campo in enumerate(self.campos_disponiveis):
            row, col = divmod(i, cols)
            cb = tk.Checkbutton(campos_frame, text=campo, variable=self.campos_selecionados[campo], bg="#f5f5f5")
            cb.grid(row=row, column=col, sticky="w", padx=5, pady=2)

        ctrl_frame = tk.Frame(main_frame, bg="#f5f5f5")
        ctrl_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Button(ctrl_frame, text="📂 Selecionar PDF", command=self.selecionar_pdf, font=("Arial", 12),
                  bg="#4caf50", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)

        tk.Button(ctrl_frame, text="⬇ Exportar para Excel", command=self.exportar_excel, font=("Arial", 12),
                  bg="#2196f3", fg="white", padx=15, pady=5).pack(side=tk.LEFT, padx=5)

        self.text_area = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, width=120, height=20,
                                                   font=("Courier New", 10), bg="#f0f0f0")
        self.text_area.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.status_bar = tk.Label(root, text="Pronto", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        self.pdf_texto = ""
        self.pdf_path = ""

    def selecionar_pdf(self):
        caminho = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
        if caminho:
            self.pdf_path = caminho
            try:
                with pdfplumber.open(caminho) as pdf:
                    texto = ""
                    for pagina in pdf.pages:
                        txt = pagina.extract_text()
                        if txt:
                            texto += txt + "\n"
                self.pdf_texto = texto
                self.text_area.delete(1.0, tk.END)
                self.text_area.insert(tk.END, self.pdf_texto)
                self.status_bar.config(text=f"PDF carregado: {caminho.split('/')[-1]}")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível ler o PDF:\n{str(e)}")
                self.status_bar.config(text="Erro ao carregar PDF")

    def exportar_excel(self):
        if not self.pdf_texto:
            messagebox.showwarning("Aviso", "Nenhum PDF carregado.")
            return

        default_name = f"dados_financeiros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=default_name)
        if not save_path:
            return

        try:
            self.status_bar.config(text="Processando dados financeiros...")
            self.root.update()

            blocos_clientes = re.split(r'-{100,}', self.pdf_texto)
            dados_clientes = []

            for bloco in blocos_clientes:
                if not bloco.strip() or "PAGINA :" in bloco or "RELACAO DE PAGAMENTOS" in bloco:
                    continue

                cliente = self.processar_bloco_cliente(bloco)
                if cliente:
                    dados_clientes.append(cliente)

            if not dados_clientes:
                messagebox.showinfo("Resultado", "Nenhum dado de cliente encontrado.")
                self.status_bar.config(text="Nenhum dado encontrado")
                return

            self.salvar_excel(save_path, dados_clientes)
            messagebox.showinfo("Sucesso", f"Arquivo '{save_path.split('/')[-1]}' salvo com sucesso!")
            self.status_bar.config(text=f"Exportação concluída: {save_path.split('/')[-1]}")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante a exportação:\n{str(e)}")
            self.status_bar.config(text="Erro durante a exportação")

    def processar_bloco_cliente(self, bloco):
        linhas = [linha.strip() for linha in bloco.split('\n') if linha.strip()]
        if len(linhas) < 2:
            return None

        cabecalho = linhas[0]

        matricula_match = re.search(r'^(\d{7})', cabecalho)
        if not matricula_match:
            return None
        matricula = matricula_match.group(1)

        nome_match = re.search(r'^\d{7}\s+\d+\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\-]+[A-ZÀ-Ú])(?=\s+\d|$)', cabecalho)
        if not nome_match:
            nome_match = re.search(r'^\d{7}\s+([A-ZÀ-Ú][A-ZÀ-Ú\s\-]+[A-ZÀ-Ú])(?=\s+\d|$)', cabecalho)
        nome = nome_match.group(1).strip() if nome_match else ""

        cpf_match = re.search(r'(\d{3}\.\d{3}\.\d{3}-\d{2}|\d{9}-\d{2})', "\n".join(linhas))
        cpf = cpf_match.group(1) if cpf_match else ""

        banco_agencia = ""
        conta_corrente = ""

        banco_conta_match = re.search(r'(\d{3,}-[A-Z0-9])\s+(\d{4,}-[\dXx\-]+)$', cabecalho)
        if not banco_conta_match and len(linhas) > 1:
            banco_conta_match = re.search(r'(\d{3,}-[A-Z0-9])\s+(\d{4,}-[\dXx\-]+)', linhas[1])

        if banco_conta_match:
            banco_agencia = banco_conta_match.group(1)
            conta_corrente = banco_conta_match.group(2)

        data_apos_match = re.search(r'(\d{2}/\d{2}/\d{4})', "\n".join(linhas[:5]))
        data_aposentadoria = data_apos_match.group(1) if data_apos_match else ""

        total_rendimentos = self.extrair_total(linhas, "TOTAL RENDIMENTOS :")
        total_descontos = self.extrair_total(linhas, "TOTAL DESCONTOS :")
        deposito_fgts = self.extrair_total(linhas, "DEPOSITO FGTS :")
        total_liquido = self.extrair_total(linhas, "TOTAL LIQUIDO :")

        margem_30 = self.extrair_margem(linhas, "MARGEM CONSIG. 30%:")
        margem_70 = self.extrair_margem(linhas, "MARGEM CONSIG. 70%:")

        descontos_bancarios = self.extrair_descontos_bancarios(linhas)

        bancos = "; ".join([d['Banco'] for d in descontos_bancarios])
        parcelas = "; ".join([d['Parcela'] for d in descontos_bancarios])
        valores = "; ".join([d['Valor'] for d in descontos_bancarios])

        return {
            "Matrícula": matricula,
            "Nome": nome,
            "CPF": cpf,
            "Banco/Agência": banco_agencia,
            "Conta Corrente": conta_corrente,
            "Data Aposentadoria": data_aposentadoria,
            "Total Rendimentos": total_rendimentos,
            "Total Descontos": total_descontos,
            "Depósito FGTS": deposito_fgts,
            "Total Líquido": total_liquido,
            "Margem Consignável 30%": margem_30,
            "Margem Consignável 70%": margem_70,
            "Bancos Desconto": bancos,
            "Parcelas Desconto": parcelas,
            "Valores Desconto": valores
        }

    def extrair_total(self, linhas, padrao):
        for linha in linhas:
            if padrao in linha:
                valor_match = re.search(rf"{re.escape(padrao)}\s*\*+([\d.,]+)", linha)
                if valor_match:
                    return valor_match.group(1)
        return "0,00"

    def extrair_margem(self, linhas, padrao):
        for linha in linhas:
            if padrao in linha:
                return linha.split(padrao)[1].strip().split()[0]
        return ""

    def extrair_descontos_bancarios(self, linhas):
        descontos = []
        padrao = re.compile(r'\d{1,2}\s+\d{5}\s+([A-Z \-]+)\s+(\d{1,3})\s+([\d.,]+)')
        for linha in linhas:
            if any(x in linha for x in ["EMPREST", "AMORT", "CONTRIB"]):
                match = padrao.search(linha)
                if match:
                    banco = match.group(1).strip()
                    parcela = match.group(2)
                    valor = match.group(3)
                    descontos.append({
                        "Banco": banco,
                        "Parcela": parcela,
                        "Valor": valor
                    })
        return descontos

    def salvar_excel(self, caminho, dados_clientes):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados Financeiros"

        cabecalhos = [campo for campo in self.campos_disponiveis if self.campos_selecionados[campo].get()]

        for col, cabecalho in enumerate(cabecalhos, start=1):
            ws.cell(row=1, column=col, value=cabecalho).font = Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

        for row, cliente in enumerate(dados_clientes, start=2):
            for col, campo in enumerate(cabecalhos, start=1):
                valor = cliente.get(campo, "")
                ws.cell(row=row, column=col, value=valor)

        wb.save(caminho)

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFExportador(root)
    root.mainloop()
